import os
# import pandas as pd
import openpyxl

from flask import Flask
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_restplus import Api, Resource, fields, abort

from sqlalchemy import Column

app = Flask(__name__)

# Configure the SQLAlchemy part of the app instance
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_ECHO'] = True
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////' + os.path.join(basedir, 'properties.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# enable CORS
CORS(app, resources={r'/*': {'origins': '*'}})

# Create the SQLAlchemy db instance
db = SQLAlchemy(app)

app_api = Api(app=app,
          version="1.0",
          title="Enodo challenge api",
          description="Allow users to search, select, or unselect properties from the DB")

api = app_api.namespace('api', description='APIs')


class Property(db.Model):
    """
    Property model. Rest of the fields will be dynamically added while server startup.
    """
    __tablename__ = 'properties'
    property_id = db.Column(db.Integer, primary_key=True)
    selected = db.Column(db.Boolean, nullable=True)


# api model to handel put request
model_with_selected = api.model('Property', {
    'property_id': fields.Integer,
    'selected': fields.Boolean
})

# api model to return the details of a property
detailed_model = api.model('Property', {
    'property_id': fields.Integer,
    'full_address': fields.String,
    'class_description': fields.String,
})


@api.route('/properties/')
class PropertyListAPI(Resource):
    @api.marshal_with(detailed_model)
    def get(self, **kwargs):
        """return all properties that are not unselected"""
        return Property.query.filter((Property.selected == None) | (Property.selected == 1)).all()


@api.route('/properties/<int:id>')
@api.response(404, 'Property not found')
@api.param('id', 'Property id')
class PropertyAPI(Resource):
    @api.expect(model_with_selected)
    @api.marshal_with(detailed_model)
    def put(self, id):
        """Update a property given its id"""
        current_property = Property.query.get(id)
        if current_property:
            current_property.selected = api.payload['selected']
            db.session.commit()
            if api.payload['selected']:
                return current_property
            else:
                return Property.query.filter((Property.selected == None) | (Property.selected == 1)).all()
        else:
            api.abort(code=404, message="Property not found.")


def build_database():
    print('Initiating app')
    # get data from excel using pandas
    # df = pd.read_excel(os.path.abspath(os.path.join(os.path.dirname(basedir), 'Enodo_Skills_Assessment_Data_File.xlsx')))
    # columns = df.columns.ravel()

    # # dynamically add columns to Property model
    # for col in columns:
    #     setattr(
    #         Property, col.lower().strip().replace(' ', '_'), Column(db.String(256), nullable=True)
    #     )
    #
    # db_file = os.path.join(basedir, 'properties.db')
    # if not os.path.exists(db_file):
    #     print('Database not found. Creating database...')
    #     # Create the database
    #     db.create_all()
    #
    #     # Iterate over the excel and populate the database
    #     for index, row in df.iterrows():
    #         property_data = {}
    #         for column in df.columns.ravel():
    #             col = column.lower().strip().replace(' ', '_')
    #             property_data[col] = row[column]
    #         p = Property(**property_data)
    #         db.session.add(p)
    #
    #     db.session.commit()

    # get data from excel using openpyxl
    # xlsx_file = os.path.abspath(os.path.join(os.path.dirname(basedir), 'Enodo_Skills_Assessment_Data_File.xlsx'))
    xlsx_file = os.path.join(basedir, 'Enodo_Skills_Assessment_Data_File.xlsx')
    obj = openpyxl.load_workbook(xlsx_file)
    sheet = obj.active

    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)

    # dynamically add columns to Property model
    for col in col_names:
        setattr(
            Property, col.lower().strip().replace(' ', '_'), Column(db.String(256), nullable=True)
        )

    db_file = os.path.join(basedir, 'properties.db')
    if not os.path.exists(db_file):
        print('Database not found. Creating database...')
        # Create the database
        db.create_all()

        # Iterate over the excel and populate the database
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                property_data = {}
                for j, column in enumerate(col_names):
                    col = column.lower().strip().replace(' ', '_')
                    property_data[col] = row[j]
                p = Property(**property_data)
                db.session.add(p)

        db.session.commit()


build_database()

if __name__ == '__main__':
    app.run()
